"""
Unimex Customs - Shipment Paperwork to Customs Summary
=====================================================
Reads one inbound shipment's paperwork and writes the customs team's Excel
file: one row per part number, aggregated across every invoice in the shipment.

Inputs (identified by CONTENT, not filename -- filenames vary by client):
  * Commercial invoices (PDF)        -- REQUIRED. Part, qty, price, PO, origin.
  * Packing list (.xls/.xlsx)        -- gross/net weight, cartons, HS groups.
  * Warehouse reception report (PDF) -- optional; independent qty cross-check.
  * Parts database (database\\*.xlsx) -- Product Key -> HTS. Not in input\\.

Output columns:
  Part # | Qty | HTS | Value | Cartons | Gross Weight (kg) | Net Weight (kg) |
  Country Origin | Charges | Zone

Run:
    python customs_processor.py                # ./input -> ./output
    python customs_processor.py /path/in /path/out

Drop one shipment's documents straight into input\\, or give each shipment its
own subfolder under input\\ to process several in one go.
"""

from __future__ import annotations

import hashlib
import math
import os
import re
import sys
import traceback
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Workbook contains no default style")

# Single source of truth for the version (see _version.py). Guarded so a stray
# copy without _version.py alongside it still runs.
try:
    from _version import CUSTOMS_VERSION as __version__
except Exception:
    __version__ = "0.0.0"

# ---------------------------------------------------------------------------
# Constants the customs team confirmed. Change here, not in the code below.
# ---------------------------------------------------------------------------
CHARGES = 3.00
ZONE = "P"

OUT_COLUMNS = [
    "Part #", "Qty", "HTS", "Value", "Cartons",
    "Gross Weight (kg)", "Net Weight (kg)", "Country Origin", "Charges", "Zone",
]

# Invoices spell the country out; the filing wants the 2-letter ISO code.
COUNTRY_ISO = {
    "CHINA": "CN", "MEXICO": "MX", "UNITED STATES": "US", "USA": "US",
    "INDIA": "IN", "THAILAND": "TH", "VIETNAM": "VN", "TAIWAN": "TW",
    "CANADA": "CA", "GERMANY": "DE", "ITALY": "IT", "SPAIN": "ES",
    "FRANCE": "FR", "JAPAN": "JP", "KOREA": "KR", "SOUTH KOREA": "KR",
    "MALAYSIA": "MY", "INDONESIA": "ID", "BRAZIL": "BR", "POLAND": "PL",
    "PORTUGAL": "PT", "TURKEY": "TR", "UNITED KINGDOM": "GB",
}

# Packing-list header keywords -> canonical column. First header whose
# normalized text contains any of these tokens wins.
PL_COLUMN_TOKENS = {
    "po":          ["purchasedorder", "purchaseorder", "po#", "pono", "p/o", "order#"],
    # 'cus.partid' first and no bare 'item': a sheet with both 'Cus. Part ID'
    # (the Kohler part) and 'Solex Item' (the supplier's own code) must pick the
    # Kohler one, since that is what the parts database is keyed on.
    "part":        ["cus.partid", "custpartid", "partid", "materialno", "partno",
                    "itemno", "material", "part#", "sku", "item#"],
    "description": ["description", "descripcion"],
    "qty":         ["qty", "quantity", "pcs", "cantidad"],
    "hs":          ["hsncode", "hscode", "hsn", "hs", "tariff", "fraccion"],
    # 'g.w'/'n.w' cover the PDF templates that abbreviate; the per-carton
    # variants ('G.W/CTN') are filtered out before these are applied.
    "gross":       ["grossweight", "gross", "g.w", "gw(", "pesobruto"],
    "net":         ["netweight", "net", "n.w", "nw(", "pesoneto"],
    "volume":      ["volume", "cbm", "volumen", "meas", "mea."],
    "cartons":     ["carton", "ctn", "cajas", "bultos", "package"],
}

# Commercial invoices that arrive as a spreadsheet rather than a PDF. Same
# keyword idea as the packing list, different columns.
INV_COLUMN_TOKENS = {
    "po":          ["po/no", "po#", "pono", "purchaseorder", "purchasedorder", "p/o"],
    "part":        ["cus.partid", "custpartid", "partid", "materialno", "partno",
                    "itemno", "material", "part#", "sku"],
    "description": ["description", "descripcion"],
    "qty":         ["qty", "quantity", "cantidad"],
    "price":       ["unitprice", "priceeach", "price"],
    "amount":      ["amount", "extendedvalue", "exttotal", "importe"],
    "hs":          ["hscode", "hsno", "tariff"],
}

# ---------------------------------------------------------------------------
# Regexes. Kept whitespace-tolerant on purpose (TRAP 6): PDF backends pad
# columns differently, so a literal space would work on one and fail on another.
# ---------------------------------------------------------------------------
# "10   24  1308865-U-BV   29.87   716.88"
RE_ITEM_LINE = re.compile(
    r"^(\d{1,4})\s+([\d,]+)\s+([A-Za-z0-9][A-Za-z0-9\-./#]*)\s+"
    r"([\d,]*\.?\d+)\s+([\d,]*\.?\d+)\s*$"
)
# The HS code the invoice prints, on its own line under the description.
RE_HS_ONLY = re.compile(r"^(\d{6,12})$")
RE_ORIGIN = re.compile(r"Country\s+of\s+Origin\s*[-:]\s*(.+?)\s*$", re.I)
RE_ITEMS_TOTAL = re.compile(r"Items?\s+total\s+([\d,]+\.\d{2})", re.I)
# Header/value pairs are not always on adjacent lines (TRAP 6): scan forward.
RE_NUM_DATE = re.compile(r"^(\S+)\s+(\d{1,2}/\d{1,2}/\d{4})\b")
# Summary page row: "19  9079663585  1030012622  1013749677BU  18,487.98"
RE_SUMMARY_ROW = re.compile(
    r"^(\d{1,4})\s+(\d{6,})\s+(\d{6,})\s+(\S+)\s+([\d,]+\.\d{2})\s*$"
)
# Reception report line, e.g.
# "CHN3609436 (L/E) 1 830915 ARANDELA DE ACERO 900.00 PIEZA 900.00 PIEZA 7318229199 S/N"
RE_RECEPTION_ROW = re.compile(
    r"\(L/E\)\s+(\d{1,4})\s+(\S+)\s+.*?([\d,]+\.\d{2})\s+([A-Za-zÁÉÍÓÚÑ]+)\s+"
    r"([\d,]+\.\d{2})\s+([A-Za-zÁÉÍÓÚÑ]+)\s+(\d{6,12})\b"
)
RE_SHIPMENT_NO = re.compile(r"shipment\s*(?:no\.?|number)", re.I)


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def base_dir() -> Path:
    """
    Folder the script (or .exe) lives in. When packaged with PyInstaller and
    double-clicked, the current working directory can be C:\\Windows\\System32,
    so we always anchor input/output/database to the .exe's actual location.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


class TeeLogger:
    """Write everything printed to console to a log file as well."""
    def __init__(self, log_path: Path):
        self.terminal = sys.stdout
        self.log = open(log_path, "w", encoding="utf-8")
    def write(self, msg: str) -> None:
        self.terminal.write(msg)
        self.log.write(msg)
        self.log.flush()
    def flush(self) -> None:
        self.terminal.flush()
        self.log.flush()
    def close(self) -> None:
        try:
            self.log.close()
        except Exception:
            pass


def pause_for_user() -> None:
    """Wait for a keypress so the console doesn't flash shut on double-click."""
    if not sys.stdout.isatty():
        return
    try:
        input("\nPress ENTER to close...")
    except (EOFError, KeyboardInterrupt):
        pass


def _num(text: object) -> float | None:
    """'1,234.56' -> 1234.56. Returns None for blanks and non-numbers."""
    if text is None:
        return None
    if isinstance(text, (int, float)) and not isinstance(text, bool):
        return None if (isinstance(text, float) and math.isnan(text)) else float(text)
    s = str(text).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _norm_key(text: object) -> str:
    """Uppercase, whitespace-stripped. The join key for part numbers and POs."""
    return "".join(str(text).split()).upper() if text is not None else ""


def _depunct(text: object) -> str:
    """Uppercase alphanumerics only -- for punctuation-blind part matching."""
    return re.sub(r"[^A-Z0-9]", "", str(text).upper()) if text is not None else ""


def _edit_distance_1(a: str, b: str) -> bool:
    """True when a and b differ by exactly one insert, delete or substitution."""
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False
    if la == lb:
        diff = sum(1 for x, y in zip(a, b) if x != y)
        return diff == 1
    if la > lb:
        a, b, la, lb = b, a, lb, la
    i = j = 0
    skipped = False
    while i < la and j < lb:
        if a[i] != b[j]:
            if skipped:
                return False
            skipped = True
            j += 1
            continue
        i += 1
        j += 1
    return True


def _iso_country(name: str) -> tuple[str, bool]:
    """('China') -> ('CN', True). Unknown names come back as-is with False so
    the caller can flag them instead of silently inventing a code."""
    raw = str(name).strip().rstrip(".").upper()
    if len(raw) == 2 and raw.isalpha():
        return raw, True
    if raw in COUNTRY_ISO:
        return COUNTRY_ISO[raw], True
    return str(name).strip(), False


def allocate_largest_remainder(total: float, weights: list[float]) -> list[float]:
    """
    Split `total` across `weights` in proportion, rounded to 2 decimals, so the
    parts sum to `total` EXACTLY. Works in integer hundredths -- floats would
    reintroduce the 530.3199999999999 artefacts the filing must not contain.
    """
    if not weights:
        return []
    grand = sum(weights)
    if grand <= 0:
        return [0.0] * len(weights)
    cents_total = int(round(total * 100))
    raw = [cents_total * w / grand for w in weights]
    floors = [int(math.floor(r)) for r in raw]
    short = cents_total - sum(floors)
    # Hand the leftover cents to the largest fractional parts, breaking ties by
    # the bigger weight so the result is deterministic run to run.
    order = sorted(range(len(raw)), key=lambda i: (-(raw[i] - floors[i]), -weights[i], i))
    for i in order[:max(short, 0)]:
        floors[i] += 1
    return [f / 100.0 for f in floors]


# ---------------------------------------------------------------------------
# PDF text extraction. Backends are NOT interchangeable (TRAP 5) -- try each
# and keep whichever actually yields text.
# ---------------------------------------------------------------------------
def extract_pdf_pages(path: Path) -> list[str]:
    """Return the text of every page, best available backend first."""
    attempts: list[str] = []
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            pages = [(p.extract_text() or "") for p in pdf.pages]
        if any(t.strip() for t in pages):
            return pages
        attempts.append("pdfplumber returned no text")
    except Exception as e:
        attempts.append(f"pdfplumber failed ({type(e).__name__})")

    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        pages = [(pg.extract_text() or "") for pg in reader.pages]
        if any(t.strip() for t in pages):
            print(f"  [INFO] {path.name}: read with pypdf ({attempts[0]}).")
            return pages
        attempts.append("pypdf returned no text")
    except Exception as e:
        attempts.append(f"pypdf failed ({type(e).__name__})")

    raise RuntimeError(f"could not read any text from {path.name}: {'; '.join(attempts)}")


# ---------------------------------------------------------------------------
# Invoices
# ---------------------------------------------------------------------------
@dataclass
class InvoiceLine:
    invoice: str
    po: str
    item: str
    part: str
    qty: float
    unit_price: float
    ext_total: float
    invoice_hs: str
    country_raw: str
    page: int


@dataclass
class SummaryEntry:
    summary_invoice: str
    item: str
    invoice: str
    sales_order: str
    po: str
    total: float


@dataclass
class InvoiceDoc:
    path: Path
    lines: list[InvoiceLine] = field(default_factory=list)
    summaries: list[SummaryEntry] = field(default_factory=list)
    summary_totals: dict[str, float] = field(default_factory=dict)
    page_total_mismatches: list[str] = field(default_factory=list)
    invoice_pages: int = 0
    summary_pages: int = 0


def _scan_forward(lines: list[str], start: int, pattern: re.Pattern, span: int = 8):
    """Find the first match of `pattern` within `span` lines after `start`.
    A header and its value are not always adjacent (TRAP 6)."""
    for line in lines[start + 1: start + 1 + span]:
        m = pattern.search(line)
        if m:
            return m
    return None


def parse_invoice_page(text: str, page_no: int) -> list[InvoiceLine]:
    """Pull every line item off one invoice page."""
    lines = [ln.rstrip() for ln in text.splitlines()]

    invoice_no = ""
    for i, ln in enumerate(lines):
        if re.search(r"\bNumber\b.*\bDate\b", ln):
            m = _scan_forward(lines, i, RE_NUM_DATE)
            if m and m.group(1).isdigit():
                invoice_no = m.group(1)
                break

    po = ""
    for i, ln in enumerate(lines):
        if re.search(r"P\.?\s*O\.?\s*No", ln, re.I):
            m = _scan_forward(lines, i, RE_NUM_DATE)
            if m:
                po = m.group(1)
                break

    # Locate the item lines first, then read each one's trailing block (the
    # description, the HS code and the country line) up to the next item.
    item_at: list[tuple[int, re.Match]] = []
    for i, ln in enumerate(lines):
        m = RE_ITEM_LINE.match(ln.strip())
        if m:
            item_at.append((i, m))

    out: list[InvoiceLine] = []
    for n, (idx, m) in enumerate(item_at):
        end = item_at[n + 1][0] if n + 1 < len(item_at) else len(lines)
        block = lines[idx + 1: end]
        hs = ""
        country = ""
        for bl in block:
            s = bl.strip()
            if not hs:
                hm = RE_HS_ONLY.match(s)
                if hm:
                    hs = hm.group(1)
            om = RE_ORIGIN.search(s)
            if om and not country:
                country = om.group(1).strip()
        out.append(InvoiceLine(
            invoice=invoice_no,
            po=po,
            item=m.group(1),
            part=_norm_key(m.group(3)),
            qty=float(m.group(2).replace(",", "")),
            unit_price=float(m.group(4).replace(",", "")),
            ext_total=float(m.group(5).replace(",", "")),
            invoice_hs=hs,
            country_raw=country,
            page=page_no,
        ))
    return out


def parse_invoices(path: Path) -> InvoiceDoc:
    """
    Read the invoice PDF. Pages titled 'Invoice Summary' are roll-ups, not
    invoices: they are collected separately for reconciliation. There may be
    MORE THAN ONE summary set in different places in the file -- every set is
    kept, keyed by its own summary invoice number.
    """
    doc = InvoiceDoc(path=path)
    pages = extract_pdf_pages(path)

    for pno, text in enumerate(pages, start=1):
        if re.search(r"Invoice\s+Summary", text, re.I):
            doc.summary_pages += 1
            lines = [ln.rstrip() for ln in text.splitlines()]
            sum_no = ""
            for i, ln in enumerate(lines):
                if re.search(r"\bNumber\b.*\bDate\b", ln):
                    m = _scan_forward(lines, i, RE_NUM_DATE)
                    if m and m.group(1).isdigit():
                        sum_no = m.group(1)
                        break
            for ln in lines:
                m = RE_SUMMARY_ROW.match(ln.strip())
                if m:
                    doc.summaries.append(SummaryEntry(
                        summary_invoice=sum_no,
                        item=m.group(1),
                        invoice=m.group(2),
                        sales_order=m.group(3),
                        po=m.group(4),
                        total=float(m.group(5).replace(",", "")),
                    ))
                tm = RE_ITEMS_TOTAL.search(ln)
                if tm and sum_no:
                    doc.summary_totals[sum_no] = float(tm.group(1).replace(",", ""))
            continue

        page_lines = parse_invoice_page(text, pno)
        if not page_lines:
            continue
        doc.invoice_pages += 1
        doc.lines.extend(page_lines)

        # Each page prints its own line-items total; check ours against it.
        printed = None
        for ln in text.splitlines():
            tm = RE_ITEMS_TOTAL.search(ln)
            if tm:
                printed = float(tm.group(1).replace(",", ""))
        if printed is not None:
            ours = round(sum(l.ext_total for l in page_lines), 2)
            if abs(ours - printed) > 0.01:
                doc.page_total_mismatches.append(
                    f"page {pno} (invoice {page_lines[0].invoice}): "
                    f"lines sum to {ours:,.2f} but page prints {printed:,.2f}")
    return doc


# ---------------------------------------------------------------------------
# Invoices, spreadsheet form
#
# Several suppliers send the commercial invoice as a worksheet rather than a
# PDF, often in the same workbook as the packing list (one sheet each), and
# sometimes as two invoice sheets covering one shipment. The table is found by
# its header keywords, so a new supplier's column order costs nothing.
# ---------------------------------------------------------------------------
def _labelled_value(df: pd.DataFrame, pattern: str, limit: int = 40) -> str:
    """Find a label like 'Invoice No:' or 'Country of Origin:' and return the
    value printed beside it (or just below), which is where these headers put
    it. Returns '' when the label is absent."""
    rx = re.compile(pattern, re.I)
    for r in range(min(limit, len(df))):
        for c in range(df.shape[1]):
            v = df.iat[r, c]
            if v is None or not rx.search(str(v)):
                continue
            for cc in range(c + 1, df.shape[1]):          # to the right
                nxt = df.iat[r, cc]
                if nxt is not None and str(nxt).strip() and str(nxt).strip().lower() != "nan":
                    return str(nxt).strip()
            for rr in range(r + 1, min(r + 3, len(df))):  # or underneath
                nxt = df.iat[rr, c]
                if nxt is not None and str(nxt).strip() and str(nxt).strip().lower() != "nan":
                    return str(nxt).strip()
            # 'MADE IN CHINA' carries the value in the same cell as the label
            m = re.search(r"made\s+in\s+([A-Za-z ]+)", str(v), re.I)
            if m:
                return m.group(1).strip()
    return ""


def parse_invoices_excel(path: Path) -> InvoiceDoc:
    """Read commercial invoice line items from a workbook. Every sheet holding
    an invoice table is read; a workbook with 'CI 1' and 'CI 2' is two invoices
    for one shipment and both belong in the same summary."""
    doc = InvoiceDoc(path=path)
    xl = pd.ExcelFile(path, engine="xlrd" if path.suffix.lower() == ".xls" else "openpyxl")

    for sheet in xl.sheet_names:
        df = _read_sheet(path, sheet)
        hit = _find_header(df, INV_COLUMN_TOKENS, {"po", "part", "qty", "amount"})
        if not hit:
            continue
        hstart, hspan, cols = hit
        # A packing sheet can also show PO/part/qty/amount-ish columns; what it
        # never has is a price per unit. Require one so the packing sheet in the
        # same workbook is not read as a second invoice.
        if "price" not in cols:
            continue

        invoice_no = _labelled_value(df, r"invoice\s*(?:no|number|#)") or path.stem
        country = _labelled_value(df, r"country\s+of\s+origin|made\s+in\s+")

        def cell(r: int, name: str):
            idx = cols.get(name)
            return df.iat[r, idx] if idx is not None and idx < df.shape[1] else None

        n_before = len(doc.lines)
        printed_total = None
        for r in range(hstart + hspan, len(df)):
            row_text = " ".join(str(v) for v in df.iloc[r] if v is not None
                                and str(v).strip().lower() != "nan")
            if re.match(r"\s*(total|grand\s*total|say\s+total)\b", row_text, re.I):
                amt = _num(cell(r, "amount"))
                if amt is not None:
                    printed_total = amt
                continue

            po = _norm_key(cell(r, "po"))
            part = _norm_key(cell(r, "part"))
            qty = _num(cell(r, "qty"))
            amount = _num(cell(r, "amount"))
            price = _num(cell(r, "price"))
            if po.lower() in {"nan", "none"}:
                po = ""
            if part.lower() in {"nan", "none"}:
                part = ""
            if not part or qty is None or amount is None:
                continue

            hs = str(cell(r, "hs") or "").strip()
            if hs.lower() in {"nan", "none"}:
                hs = ""
            hs = re.sub(r"\.0+$", "", hs)   # '84819090.00' is a code, not a number
            doc.lines.append(InvoiceLine(
                invoice=invoice_no, po=po, item=str(r),
                part=part, qty=qty,
                unit_price=price if price is not None else (amount / qty if qty else 0.0),
                ext_total=amount, invoice_hs=hs,
                country_raw=country, page=xl.sheet_names.index(sheet) + 1,
            ))

        if len(doc.lines) > n_before:
            doc.invoice_pages += 1
            if printed_total is not None:
                ours = round(sum(l.ext_total for l in doc.lines[n_before:]), 2)
                if abs(ours - printed_total) > 0.02:
                    doc.page_total_mismatches.append(
                        f"sheet '{sheet}' (invoice {invoice_no}): lines sum to "
                        f"{ours:,.2f} but the sheet's total says {printed_total:,.2f}")
                else:
                    # The sheet's own total is this document's roll-up; record it
                    # so the summary check has something to verify against.
                    doc.summary_totals[f"{invoice_no}/{sheet}"] = printed_total

    if not doc.lines:
        raise RuntimeError(f"{path.name}: no invoice line items found "
                           f"(needs PO, part, qty, unit price and amount columns)")
    return doc


# ---------------------------------------------------------------------------
# Packing list
# ---------------------------------------------------------------------------
@dataclass
class PackingRow:
    po: str
    part: str
    qty: float
    hs: str
    group: int
    cartons: float | None = None   # only some templates state cartons per row
    consumed: bool = False


@dataclass
class PackingList:
    path: Path
    rows: list[PackingRow] = field(default_factory=list)
    group_gross: dict[int, float] = field(default_factory=dict)
    group_net: dict[int, float] = field(default_factory=dict)
    total_qty: float | None = None
    total_gross: float | None = None
    total_net: float | None = None
    cartons: float | None = None
    carton_source: str = ""
    shipment_no: str = ""
    notes: list[str] = field(default_factory=list)


def _read_sheet(path: Path, sheet) -> pd.DataFrame:
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    return pd.read_excel(path, sheet_name=sheet, header=None, dtype=object, engine=engine)


def _map_columns(header_row: list[object], token_map: dict[str, list[str]]) -> dict[str, int]:
    """Map canonical names to column indexes by keyword. Tokens are tried in the
    order listed, most specific first, so 'grossweight' can't be claimed by the
    bare 'gross' rule and 'Solex Item' can't be claimed ahead of 'Cus. Part ID'.
    Per-carton columns ('N.W/CTN') are never eligible: they are per-unit figures,
    not row totals."""
    normed = ["".join(str(h).lower().split()) if h is not None else "" for h in header_row]
    found: dict[str, int] = {}
    for canon, tokens in token_map.items():
        for tok in tokens:
            candidates = [idx for idx, h in enumerate(normed)
                          if h and idx not in found.values()
                          and not any(m in h for m in PL_PER_UNIT_MARKERS)
                          and tok in h]
            if not candidates:
                continue
            # The best match is the column whose header is CLOSEST to the token,
            # not the leftmost one containing it. One supplier leaves a stray
            # 'PO#' label above an unrelated column; taking it in preference to
            # the real 'PO #' header mapped the PO onto an empty column and
            # every join then failed.
            found[canon] = min(candidates, key=lambda i: (len(normed[i]), i))
            break
    return found


def _map_pl_columns(header_row: list[object]) -> dict[str, int]:
    return _map_columns(header_row, PL_COLUMN_TOKENS)


def _merge_header_rows(df: pd.DataFrame, start: int, span: int) -> list[object]:
    """Join `span` consecutive rows into one header line, per column.

    A spreadsheet header is often stacked: 'QTY' over '(PCS)', or 'N.W.' over
    '(KGS)', and one template puts 'HS CODE' two rows below its neighbours.
    Reading a single row would leave those columns unnamed."""
    out: list[object] = []
    for c in range(df.shape[1]):
        parts = []
        for r in range(start, min(start + span, len(df))):
            v = df.iat[r, c]
            if v is not None and str(v).strip() and str(v).strip().lower() != "nan":
                parts.append(str(v).strip())
        out.append(" ".join(parts))
    return out


def _find_header(df: pd.DataFrame, token_map: dict[str, list[str]],
                 required: set[str], scan: int = 40) -> tuple[int, int, dict[str, int]] | None:
    """Locate a table header, allowing it to span up to 3 rows. Returns
    (first row, rows used, column map) for the mapping that names the most
    columns -- more named columns means the header was read correctly."""
    best: tuple[int, int, int, dict[str, int]] | None = None
    for r in range(min(scan, len(df))):
        for span in (1, 2, 3):
            cols = _map_columns(_merge_header_rows(df, r, span), token_map)
            if not required <= set(cols):
                continue
            score = len(cols)
            if best is None or score > best[0]:
                best = (score, r, span, cols)
    if best is None:
        return None
    return best[1], best[2], best[3]


def parse_packing_list(path: Path) -> PackingList:
    """
    Read the packing list. Two layout habits matter and both are common:
      * gross/net weight is stated ONCE per HS group, on the group's first row
        and carried down implicitly -- so a non-blank gross starts a new group;
      * cartons/volume are stated once for the WHOLE shipment, not per line.
    """
    pl = PackingList(path=path)
    xl = pd.ExcelFile(path, engine="xlrd" if path.suffix.lower() == ".xls" else "openpyxl")

    # Read EVERY packing sheet in the workbook. One supplier splits a shipment
    # across 'PK40GP1' and 'PK40GP2'; reading only the best-scoring sheet would
    # silently drop half the shipment's weights.
    sheets: list[tuple[str, pd.DataFrame, int, int, dict[str, int]]] = []
    for sheet in xl.sheet_names:
        df = _read_sheet(path, sheet)
        hit = _find_header(df, PL_COLUMN_TOKENS, {"po", "part", "qty"})
        # PO + part + qty alone describes an invoice table just as well as a
        # packing one. What makes it a packing list is weights or cartons;
        # without them an invoice sheet gets read as a packing list with no
        # weights, which silently suppresses the "no packing list" warning.
        if not hit or not ({"gross", "net", "cartons"} & set(hit[2])):
            continue
        # An invoice sheet in the same workbook often carries a cartons column
        # too. What it has and a packing sheet never does is a unit price --
        # without this test the invoice sheet is read as a second packing sheet
        # and every carton and weight is counted twice.
        inv = _find_header(df, INV_COLUMN_TOKENS, {"po", "part", "qty", "amount"})
        if inv and "price" in inv[2]:
            continue
        sheets.append((sheet, df, hit[0], hit[1], hit[2]))
    if not sheets:
        raise RuntimeError(f"{path.name}: could not find a packing-list header row "
                           f"(need Purchase Order / Material / Qty columns, plus "
                           f"weights or cartons)")

    group = -1
    carton_values: list[float] = []
    for sheet, df, hstart, hspan, cols in sheets:
        hrow = hstart + hspan - 1
        group, extra = _read_packing_sheet(pl, df, hrow, cols, group)
        carton_values.extend(extra)

    if pl.cartons is None and carton_values:
        distinct = {round(v, 4) for v in carton_values}
        if len(distinct) == 1 and len(carton_values) < len(pl.rows):
            # One number for the whole shipment, printed on the first row --
            # summing the column here would double-count it.
            pl.cartons = carton_values[0]
            pl.carton_source = "single shipment-wide carton figure"
        else:
            pl.cartons = sum(carton_values)
            pl.carton_source = "sum of the per-row carton column"

    # An HS group whose code is not constant means the group boundaries (drawn
    # from where weights are stated) don't line up with the tariff grouping.
    for g in sorted(pl.group_gross):
        codes = {r.hs for r in pl.rows if r.group == g and r.hs}
        if len(codes) > 1:
            pl.notes.append(f"weight group {g + 1} spans more than one HS code: "
                            f"{', '.join(sorted(codes))}")
    return pl


def _read_packing_sheet(pl: PackingList, df: pd.DataFrame, hrow: int,
                        cols: dict[str, int], group: int) -> tuple[int, list[float]]:
    """Read one packing sheet's rows into `pl`, continuing the weight-group
    numbering from `group`. Returns the new group counter and the carton
    figures seen, which the caller needs to decide how to read the column."""
    # Shipment number, printed above the table.
    for r in range(hrow + 1):
        for c in range(df.shape[1]):
            if RE_SHIPMENT_NO.search(str(df.iat[r, c])):
                for rr in range(r + 1, min(r + 4, len(df))):
                    v = df.iat[rr, c]
                    if v is not None and str(v).strip() and str(v).strip().lower() != "nan":
                        pl.shipment_no = str(v).strip().split(".")[0]
                        break
            if pl.shipment_no:
                break
        if pl.shipment_no:
            break

    def cell(r: int, name: str):
        idx = cols.get(name)
        return df.iat[r, idx] if idx is not None and idx < df.shape[1] else None

    carton_values: list[float] = []
    for r in range(hrow + 1, len(df)):
        po = _norm_key(cell(r, "po")) if cell(r, "po") is not None else ""
        part = _norm_key(cell(r, "part")) if cell(r, "part") is not None else ""
        qty = _num(cell(r, "qty"))
        gross = _num(cell(r, "gross"))
        net = _num(cell(r, "net"))
        ctn = _num(cell(r, "cartons"))
        po = "" if po.lower() in {"nan", "none"} else po
        part = "" if part.lower() in {"nan", "none"} else part

        # A totals row can also announce itself in words -- 'TOTAL' printed in
        # the PO column, or in a label column to its left. Without this the row
        # is read as a line item for a part called 'TOTAL'.
        row_text = " ".join(str(v) for v in df.iloc[r] if v is not None
                            and str(v).strip().lower() != "nan")
        if re.match(r"\s*(total|grand\s*total|say\s+total)\b", row_text, re.I):
            if qty is not None:
                # Accumulate: a shipment split across two packing sheets prints
                # a total on each, and only their sum describes the shipment.
                pl.total_qty = (pl.total_qty or 0.0) + qty
                if gross is not None:
                    pl.total_gross = (pl.total_gross or 0.0) + gross
                if net is not None:
                    pl.total_net = (pl.total_net or 0.0) + net
                if ctn is not None:
                    pl.cartons = (pl.cartons or 0.0) + ctn
                    pl.carton_source = "packing-list total row"
            continue

        if not po and not part:
            # Trailing totals row: no PO, no part, but the columns still add up.
            if qty is not None:
                # Accumulate: a shipment split across two packing sheets prints
                # a total on each, and only their sum describes the shipment.
                pl.total_qty = (pl.total_qty or 0.0) + qty
                if gross is not None:
                    pl.total_gross = (pl.total_gross or 0.0) + gross
                if net is not None:
                    pl.total_net = (pl.total_net or 0.0) + net
                if ctn is not None:
                    pl.cartons = (pl.cartons or 0.0) + ctn
                    pl.carton_source = "packing-list total row"
            continue
        if qty is None:
            continue

        if gross is not None:
            group += 1
            pl.group_gross[group] = gross
            pl.group_net[group] = net
        if group < 0:  # rows before any stated weight
            group = 0
            pl.group_gross.setdefault(group, 0.0)
            pl.group_net.setdefault(group, None)
        if ctn is not None:
            carton_values.append(ctn)

        pl.rows.append(PackingRow(
            po=po, part=part, qty=qty,
            hs=str(cell(r, "hs") or "").strip(), group=group, cartons=ctn,
        ))

    return group, carton_values


# ---------------------------------------------------------------------------
# Packing list, PDF form
#
# PDF packing lists carry no column structure -- only ink at coordinates -- and
# every supplier's template differs. Rather than a regex per supplier, the table
# is rebuilt geometrically: find the header, take each header cell's horizontal
# span as a column, then assign every word below it to the column it sits under.
#
# The parse is only trusted if it reproduces the document's own printed totals
# (section "validate"). A packing list that does not tie out is REFUSED, and the
# caller falls back to running without one -- blank weights and a loud warning --
# because wrong weights in a filing are far worse than absent ones.
# ---------------------------------------------------------------------------
# Header cells naming a PER-CARTON figure ("N.W/CTN", "Pcs/ctn", "Vol./ctn").
# These must never be mistaken for the row's total weight or quantity.
PL_PER_UNIT_MARKERS = ("/ctn", "/ ctn", "perctn", "/carton", "/box", "percarton")

# Horizontal gap, in points, below which two header words are taken to be part
# of the same header cell rather than two columns. A word space is ~2-3pt here;
# the narrowest real gutter observed is 8pt.
HEADER_WORD_GAP = 6.0


def _money(text: object) -> float | None:
    """'US$12,150.00' -> 12150.0. Currency marks and thousands separators are
    presentation; the number underneath is what matters."""
    if text is None:
        return None
    s = re.sub(r"[^\d.\-]", "", str(text).replace(",", ""))
    if not s or s in {".", "-"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _lead_num(text: object) -> float | None:
    """Leading number of a cell: '3,240 pcs' -> 3240.0, '216CTNS' -> 216.0."""
    if text is None:
        return None
    m = re.match(r"\s*([\d,]+(?:\.\d+)?)", str(text))
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None


def _cluster_lines(words: list[dict], ytol: float = 2.5) -> list[list[dict]]:
    """Group words into visual lines by their vertical position."""
    lines: dict[int, list[dict]] = {}
    for w in words:
        lines.setdefault(int(round(w["top"] / ytol)), []).append(w)
    return [sorted(ws, key=lambda w: w["x0"]) for _, ws in sorted(lines.items())]


def _header_columns(band: list[list[dict]]) -> list[dict]:
    """
    Turn the header band's words into columns.

    A header cell can be stacked over several physical lines --
    'N.W/CT' / 'N' / '(KGS)' are one column -- so words are clustered by
    HORIZONTAL OVERLAP across the whole band, not by line.
    """
    flat = [w for line in band for w in line]
    if not flat:
        return []
    cols: list[dict] = []
    for w in sorted(flat, key=lambda w: w["x0"]):
        hit = next((c for c in cols if w["x0"] <= c["x1"] + 1 and w["x1"] >= c["x0"] - 1), None)
        if hit:
            hit["x0"] = min(hit["x0"], w["x0"])
            hit["x1"] = max(hit["x1"], w["x1"])
            hit["words"].append(w)
        else:
            cols.append({"x0": w["x0"], "x1": w["x1"], "words": [w]})

    # A header cell's own words are separated by a normal word space
    # ('Kohler PO#', 'DESCRIPTION OF GOODS'); the gutter between two columns is
    # several times wider. Merge across the small gaps so one cell stays one
    # column, or the values underneath land in a column with no name.
    merged: list[dict] = []
    for c in sorted(cols, key=lambda c: c["x0"]):
        if merged and c["x0"] - merged[-1]["x1"] <= HEADER_WORD_GAP:
            merged[-1]["x1"] = max(merged[-1]["x1"], c["x1"])
            merged[-1]["words"].extend(c["words"])
        else:
            merged.append(c)
    cols = merged
    for c in cols:
        c["words"].sort(key=lambda w: (w["top"], w["x0"]))
        c["text"] = " ".join(w["text"] for w in c["words"])
        c["norm"] = "".join(c["text"].lower().split())
    return sorted(cols, key=lambda c: c["x0"])


def _assign_columns(cols: list[dict], line: list[dict]) -> dict[int, str]:
    """Bucket a data line's words into columns by where they sit horizontally.
    Values are often wider than their header, so a word that overhangs every
    column falls to the nearest one by centre distance."""
    out: dict[int, list[str]] = {}
    for w in line:
        mid = (w["x0"] + w["x1"]) / 2
        idx = next((i for i, c in enumerate(cols) if c["x0"] - 1 <= mid <= c["x1"] + 1), None)
        if idx is None:
            idx = min(range(len(cols)),
                      key=lambda i: abs(mid - (cols[i]["x0"] + cols[i]["x1"]) / 2))
        out.setdefault(idx, []).append(w["text"])
    return {i: " ".join(v) for i, v in out.items()}


def _map_pdf_columns(cols: list[dict]) -> dict[str, int]:
    """Map canonical names onto header columns, per-carton columns excluded."""
    found: dict[str, int] = {}
    eligible = [i for i, c in enumerate(cols)
                if not any(m in c["norm"] for m in PL_PER_UNIT_MARKERS)]
    for canon, tokens in PL_COLUMN_TOKENS.items():
        for tok in tokens:
            for i in eligible:
                if i in found.values():
                    continue
                if tok in cols[i]["norm"]:
                    found[canon] = i
                    break
            if canon in found:
                break
    return found


def _part_from_text(text: str) -> str:
    """Templates without a part column print it as the first token of the
    description: '1195792-CP TRIP LEVER ASSEMBLY' -> '1195792-CP'."""
    for tok in str(text).split():
        cleaned = tok.strip(",;:()")
        if len(cleaned) >= 4 and any(ch.isdigit() for ch in cleaned) \
                and re.fullmatch(r"[0-9A-Za-z][0-9A-Za-z\-./]*", cleaned):
            return _norm_key(cleaned)
    return ""


def parse_packing_list_pdf(path: Path) -> PackingList:
    """Read a packing list that arrived as a PDF. Raises if it cannot be
    parsed and validated -- never returns a half-read table."""
    import pdfplumber

    pl = PackingList(path=path)
    header_cols: list[dict] = []
    cols_map: dict[str, int] = {}
    totals_line: dict[int, str] = {}
    group = -1

    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            lines = _cluster_lines(page.extract_words())
            if not lines:
                continue

            # Locate the header on this page: the run of up to 5 consecutive
            # lines that together name the most canonical columns. Continuation
            # pages often repeat it; if this page has none, keep the last one.
            best: tuple[int, list[dict], dict[str, int]] | None = None
            for start in range(min(len(lines), 60)):
                for span in range(1, 6):
                    if start + span > len(lines):
                        break
                    cols = _header_columns(lines[start:start + span])
                    mapped = _map_pdf_columns(cols)
                    score = len(mapped) + (2 if {"qty", "gross"} <= set(mapped) else 0)
                    if {"qty"} <= set(mapped) and (best is None or score > best[0]):
                        best = (score, cols, mapped)
            if best and best[0] >= 4:
                header_cols, cols_map = best[1], best[2]
                header_bottom = max(w["bottom"] for c in header_cols for w in c["words"])
            elif header_cols:
                header_bottom = 0.0
            else:
                continue

            def val(cells: dict[int, str], name: str) -> str:
                idx = cols_map.get(name)
                return cells.get(idx, "") if idx is not None else ""

            # Pass 1: split the body into rows that carry numbers and lines that
            # carry only text. A line item's description often spans several
            # printed lines, above AND below its figures.
            data: list[tuple[float, dict[int, str]]] = []
            loose: list[tuple[float, str]] = []
            for line in lines:
                if line[0]["top"] <= header_bottom:
                    continue
                cells = _assign_columns(header_cols, line)
                text = " ".join(w["text"] for w in line)
                if re.match(r"\s*total", text, re.I):
                    for k, v in cells.items():
                        totals_line.setdefault(k, v)
                    continue
                if _lead_num(val(cells, "qty")) is None:
                    desc = val(cells, "description") or text
                    loose.append((line[0]["top"], desc))
                else:
                    data.append((line[0]["top"], cells))

            # Pass 2: every text-only line belongs to the row it sits closest
            # to vertically -- which is how a person reads it off the page.
            extra: dict[int, list[str]] = {}
            for top, desc in loose:
                if not data:
                    break
                i = min(range(len(data)), key=lambda k: abs(data[k][0] - top))
                extra.setdefault(i, []).append(desc)

            for i, (_, cells) in enumerate(data):
                qty = _lead_num(val(cells, "qty"))
                gross = _lead_num(val(cells, "gross"))
                net = _lead_num(val(cells, "net"))

                description = " ".join([val(cells, "description")] + extra.get(i, [])).strip()
                part = _norm_key(val(cells, "part")) or _part_from_text(description)
                if not part:
                    continue
                if gross is not None:
                    group += 1
                    pl.group_gross[group] = gross
                    pl.group_net[group] = net
                elif group < 0:
                    group = 0
                    pl.group_gross.setdefault(group, 0.0)
                    pl.group_net.setdefault(group, None)

                pl.rows.append(PackingRow(
                    po=_norm_key(val(cells, "po")),
                    part=part,
                    qty=qty,
                    hs=re.sub(r"\D", "", val(cells, "hs")),
                    group=group,
                    cartons=_lead_num(val(cells, "cartons")),
                ))

    if not pl.rows:
        raise RuntimeError(f"{path.name}: looks like a packing list but no line "
                           f"items could be read from it")
    if not cols_map:
        raise RuntimeError(f"{path.name}: no packing-list header row found")

    def total_of(name: str) -> float | None:
        idx = cols_map.get(name)
        return _lead_num(totals_line.get(idx)) if idx is not None else None

    # ---- validate: the parse must reproduce the document's own totals -------
    #
    # Gather every total the document states, from BOTH its TOTAL row and any
    # prose restatement ("TOTAL SAYS 262CTNS ,G.W 2648.7KGS"). Both are needed:
    # if a column was mis-identified, that column's total row cell is empty, so
    # checking only the column total would skip the check on exactly the parse
    # that got it wrong -- a missing value passing as a match.
    stated: dict[str, list[float]] = {"qty": [], "gross": [], "net": [], "cartons": []}
    for name in stated:
        v = total_of(name)
        if v is not None:
            stated[name].append(v)

    prose = {
        "gross":   r"(?:G\.?\s*W\.?|GROSS\s*WEIGHT)[^0-9]{0,12}([\d,]+\.?\d*)",
        "net":     r"(?:N\.?\s*W\.?|NET\s*WEIGHT)[^0-9]{0,12}([\d,]+\.?\d*)",
        "cartons": r"([\d,]+)\s*(?:CTNS?|CARTONS?)\b",
    }
    for page_text in extract_pdf_pages(path):
        for ln in page_text.splitlines():
            if not re.search(r"total", ln, re.I):
                continue
            for name, pattern in prose.items():
                for m in re.finditer(pattern, ln, re.I):
                    val_ = _lead_num(m.group(1))
                    if val_ is not None:
                        stated[name].append(val_)

    def check(label: str, ours: float, values: list[float], tol: float) -> None:
        for v in values:
            if abs(ours - v) > tol:
                raise RuntimeError(
                    f"{path.name}: parsed {label} {ours:,.2f} does not match the "
                    f"{v:,.2f} stated on the document -- refusing this packing list "
                    f"rather than filing figures that may be wrong")

    row_cartons = [r.cartons for r in pl.rows if r.cartons is not None]
    check("quantity", sum(r.qty for r in pl.rows), stated["qty"], 0.5)
    check("gross weight", sum(pl.group_gross.values()), stated["gross"], 0.05)
    check("net weight", sum(v for v in pl.group_net.values() if v is not None),
          stated["net"], 0.05)
    if len(row_cartons) == len(pl.rows) and row_cartons:
        check("carton count", sum(row_cartons), stated["cartons"], 0.5)

    pl.total_qty = stated["qty"][0] if stated["qty"] else None
    pl.total_gross = stated["gross"][0] if stated["gross"] else None
    pl.total_net = stated["net"][0] if stated["net"] else None

    if len(row_cartons) == len(pl.rows) and row_cartons:
        pl.cartons = sum(row_cartons)
        pl.carton_source = "per-row carton counts"
    elif stated["cartons"]:
        pl.cartons = stated["cartons"][0]
        pl.carton_source = "the packing list's stated total"

    # Say so when there was nothing to check against, rather than letting an
    # uncorroborated parse look as solid as a verified one.
    for name, label in (("gross", "gross weight"), ("net", "net weight")):
        if not stated[name]:
            pl.notes.append(f"the packing list states no total {label}, so the figures "
                            f"read from it could not be corroborated against the document")

    m = re.search(r"(?:shipment|s/?o|booking)\s*(?:no\.?|number|#)?\s*[:.]?\s*(\S+)",
                  "\n".join(extract_pdf_pages(path)[:1]), re.I)
    if m and re.search(r"\d", m.group(1)):
        pl.shipment_no = m.group(1).strip(":#.,")
    return pl


def parse_invoices_pdf_table(path: Path) -> InvoiceDoc:
    """
    Read a PDF invoice laid out as a table rather than as the classic
    one-invoice-per-page form -- same geometric approach as the PDF packing
    list, and often the very same supplier template with different columns.
    """
    import pdfplumber

    doc = InvoiceDoc(path=path)
    header_cols: list[dict] = []
    cols_map: dict[str, int] = {}
    invoice_no = ""
    country = ""

    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            lines = _cluster_lines(page.extract_words())
            if not lines:
                continue
            text = "\n".join(" ".join(w["text"] for w in ln) for ln in lines)
            if not invoice_no:
                m = re.search(r"invoice\s*(?:no|number|#)\.?\s*[:.]?\s*(\S+)", text, re.I)
                if m:
                    invoice_no = m.group(1).strip(":#.,")
            if not country:
                m = re.search(r"(?:country\s+of\s+origin|made\s+in)\s*[:.]?\s*([A-Za-z ]+)",
                              text, re.I)
                if m:
                    country = m.group(1).strip()

            best = None
            for start in range(min(len(lines), 60)):
                for span in range(1, 5):
                    if start + span > len(lines):
                        break
                    cols = _header_columns(lines[start:start + span])
                    mapped = _map_columns([c["text"] for c in cols], INV_COLUMN_TOKENS)
                    if {"qty", "amount"} <= set(mapped) and (best is None
                                                             or len(mapped) > best[0]):
                        best = (len(mapped), cols, mapped)
            if best and best[0] >= 3:
                header_cols, cols_map = best[1], best[2]
                header_bottom = max(w["bottom"] for c in header_cols for w in c["words"])
            elif header_cols:
                header_bottom = 0.0
            else:
                continue

            def val(cells: dict[int, str], name: str) -> str:
                idx = cols_map.get(name)
                return cells.get(idx, "") if idx is not None else ""

            data: list[tuple[float, dict[int, str]]] = []
            loose: list[tuple[float, str]] = []
            for line in lines:
                if line[0]["top"] <= header_bottom:
                    continue
                cells = _assign_columns(header_cols, line)
                row_text = " ".join(w["text"] for w in line)
                if re.match(r"\s*(total|grand\s*total|say\s+total)", row_text, re.I):
                    continue
                if _money(val(cells, "amount")) is None or _lead_num(val(cells, "qty")) is None:
                    loose.append((line[0]["top"], val(cells, "description") or row_text))
                else:
                    data.append((line[0]["top"], cells))

            extra: dict[int, list[str]] = {}
            for top, desc in loose:
                if not data:
                    break
                i = min(range(len(data)), key=lambda k: abs(data[k][0] - top))
                extra.setdefault(i, []).append(desc)

            for i, (_, cells) in enumerate(data):
                description = " ".join([val(cells, "description")] + extra.get(i, [])).strip()
                part = _norm_key(val(cells, "part")) or _part_from_text(description)
                qty = _lead_num(val(cells, "qty"))
                amount = _money(val(cells, "amount"))
                price = _money(val(cells, "price"))
                if not part or qty is None or amount is None:
                    continue
                doc.lines.append(InvoiceLine(
                    invoice=invoice_no or path.stem, po=_norm_key(val(cells, "po")),
                    item=str(i + 1), part=part, qty=qty,
                    unit_price=price if price is not None else (amount / qty if qty else 0.0),
                    ext_total=amount, invoice_hs=re.sub(r"\D", "", val(cells, "hs")),
                    country_raw=country, page=1,
                ))
            if doc.lines:
                doc.invoice_pages += 1

    if not doc.lines:
        raise RuntimeError(f"{path.name}: no invoice line items could be read")
    return doc


# ---------------------------------------------------------------------------
# Reception report (optional cross-check)
# ---------------------------------------------------------------------------
@dataclass
class ReceptionRow:
    partida: str
    part: str
    qty_doc: float
    qty_physical: float
    fraccion: str


@dataclass
class ReceptionReport:
    path: Path
    rows: list[ReceptionRow] = field(default_factory=list)
    gross_weight: float | None = None
    packages: str = ""
    guide: str = ""


def _glyph_chains(chars: list[dict], tol: float = 0.6) -> list[tuple[str, float]]:
    """
    Rebuild one printed row's cells by following CONTIGUOUS GLYPH RUNS.

    Two columns can physically overlap in the page layout -- on this reception
    report the tariff column runs to x=698 while the purchase-order column
    begins at x=692. Every text extractor sorts glyphs left to right and
    interleaves them:
        7419809999 + 1013878363BU  ->  74198099919013878363BU
    whose first ten characters are 7419809991: wrong, but still shaped like a
    valid code, so it passes unnoticed (TRAP 1).

    Sorting by x and chaining each character onto the one that ENDS where it
    STARTS separates the two columns cleanly, because a character of the tariff
    column continues the tariff column's chain, not the PO's.

    Returns [(text, x0)] in left-to-right order of each chain's start.
    """
    if not chars:
        return []
    ordered = sorted(chars, key=lambda c: c["x0"])
    runs: list[list[dict]] = []
    cur = [ordered[0]]
    for prev, nxt in zip(ordered, ordered[1:]):
        if abs(nxt["x0"] - prev["x1"]) <= tol:
            cur.append(nxt)
        else:
            runs.append(cur)
            cur = [nxt]
    runs.append(cur)

    # A run that starts exactly where an earlier run ended continues that same
    # chain -- this is the step that reunites '848190050' with its final '3'
    # across the interleaved character from the neighbouring column.
    chains: list[list[dict]] = []
    used = [False] * len(runs)
    for i, run in enumerate(runs):
        if used[i]:
            continue
        used[i] = True
        chain = list(run)
        extended = True
        while extended:
            extended = False
            for j, other in enumerate(runs):
                if used[j]:
                    continue
                if abs(other[0]["x0"] - chain[-1]["x1"]) <= tol:
                    chain.extend(other)
                    used[j] = True
                    extended = True
        chains.append(chain)
    chains.sort(key=lambda ch: ch[0]["x0"])
    return [("".join(c["text"] for c in ch).strip(), ch[0]["x0"]) for ch in chains]


def _reception_rows_by_glyph(path: Path) -> list[ReceptionRow]:
    """Read the report's line items cell by cell using glyph runs."""
    import pdfplumber

    rows: list[ReceptionRow] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            by_line: dict[float, list[dict]] = {}
            for ch in page.chars:
                by_line.setdefault(round(ch["top"], 0), []).append(ch)
            for top in sorted(by_line):
                cells = _glyph_chains(by_line[top])
                texts = [t for t, _ in cells]
                marker = next((i for i, t in enumerate(texts) if "(L/E)" in t), None)
                if marker is None:
                    continue
                rest = texts[marker + 1:]
                if len(rest) < 3 or not rest[0].isdigit():
                    continue
                qtys = [t for t in rest if re.fullmatch(r"[\d,]+\.\d{1,3}", t)]
                if not qtys:
                    continue
                # The tariff column is printed before the purchase-order column,
                # so the FIRST bare digit run after the quantities is the code.
                after_qty = rest[rest.index(qtys[-1]) + 1:]
                frac = next((t for t in after_qty if re.fullmatch(r"\d{6,12}", t)), "")
                rows.append(ReceptionRow(
                    partida=rest[0],
                    part=_norm_key(rest[1]),
                    qty_doc=float(qtys[0].replace(",", "")),
                    qty_physical=float(qtys[-1].replace(",", "")),
                    fraccion=frac,
                ))
    return rows


def parse_reception_report(path: Path) -> ReceptionReport:
    """
    Parse the warehouse intake report. Since the HTS now comes from the parts
    database, this document is only a cross-check: part numbers and quantities.
    The fraccion is still read so a Mexican-vs-US code difference can be
    reported, never to fill a cell.
    """
    rep = ReceptionReport(path=path)
    try:
        rep.rows = _reception_rows_by_glyph(path)
    except Exception as e:
        print(f"  [INFO] {path.name}: glyph reconstruction unavailable "
              f"({type(e).__name__}); falling back to flat text.")

    pages = extract_pdf_pages(path)
    if not rep.rows:
        # Flat-text fallback. Rows whose tariff column overlaps the PO column
        # simply will not match here -- that is deliberate. A row is dropped
        # rather than recorded with a mangled code.
        for text in pages:
            for ln in text.splitlines():
                m = RE_RECEPTION_ROW.search(ln)
                if m:
                    rep.rows.append(ReceptionRow(
                        partida=m.group(1),
                        part=_norm_key(m.group(2)),
                        qty_doc=float(m.group(3).replace(",", "")),
                        qty_physical=float(m.group(5).replace(",", "")),
                        fraccion=m.group(7),
                    ))

    for text in pages:
        for ln in text.splitlines():
            gm = re.search(r"Peso\s+bruto\s+([\d,]+\.?\d*)", ln, re.I)
            if gm and rep.gross_weight is None:
                rep.gross_weight = float(gm.group(1).replace(",", ""))
            bm = re.search(r"Bultos\s+(.+)$", ln, re.I)
            if bm and not rep.packages:
                rep.packages = bm.group(1).strip()
            gu = re.search(r"Gu[ií]?[óo]?a\(s\)\s+(\S+)", ln, re.I)
            if gu and not rep.guide:
                rep.guide = gu.group(1).strip()

    # Every code should come out the same length; a short one means a chain was
    # cut and the value cannot be trusted.
    odd = {r.fraccion for r in rep.rows if r.fraccion and not (8 <= len(r.fraccion) <= 12)}
    if odd:
        print(f"  [WARN] {path.name}: implausible tariff code(s) read from the report: "
              f"{', '.join(sorted(odd))}")
    return rep


# ---------------------------------------------------------------------------
# Parts database  ->  HTS
# ---------------------------------------------------------------------------
@dataclass
class PartsDB:
    path: Path
    by_key: dict[str, str] = field(default_factory=dict)          # exact key -> HTS
    by_depunct: dict[str, set] = field(default_factory=dict)      # A1B2 -> {HTS}
    by_stem: dict[str, set] = field(default_factory=dict)         # key minus finish -> {HTS}
    by_base: dict[str, set] = field(default_factory=dict)         # leading segment -> {HTS}


def _segments(key: str) -> list[str]:
    return [s for s in re.split(r"[-\s]+", key.strip().upper()) if s]


def load_parts_db(db_dir: Path) -> PartsDB:
    """
    Load the newest Product Key -> HTS workbook from database\\.

    Newest = most recently modified, so dropping in a fresh export is all the
    team has to do; nothing needs rebuilding or renaming.
    """
    if not db_dir.exists():
        raise RuntimeError(f"no database folder at {db_dir}")
    candidates = [p for p in db_dir.iterdir()
                  if p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
                  and not p.name.startswith("~$")]
    if not candidates:
        raise RuntimeError(f"no parts database found in {db_dir} "
                           f"(expected an .xlsx with Product Key and HTS columns)")
    path = max(candidates, key=lambda p: p.stat().st_mtime)

    df = pd.read_excel(path, dtype=str,
                       engine="xlrd" if path.suffix.lower() == ".xls" else "openpyxl")
    normed = {"".join(str(c).lower().split()): c for c in df.columns}
    key_col = next((normed[k] for k in normed if "productkey" in k or k in {"part", "partno", "sku"}), None)
    hts_col = next((normed[k] for k in normed if k.startswith("hts")), None)
    if key_col is None or hts_col is None:
        raise RuntimeError(f"{path.name}: expected 'Product Key' and 'HTS' columns, "
                           f"found {list(df.columns)[:6]}")

    db = PartsDB(path=path)
    for key, hts in zip(df[key_col], df[hts_col]):
        if key is None or hts is None:
            continue
        k = _norm_key(key)
        h = str(hts).strip()
        if not k or not h or h.lower() == "nan":
            continue
        db.by_key[k] = h
        db.by_depunct.setdefault(_depunct(k), set()).add(h)
        segs = _segments(k)
        if segs:
            db.by_base.setdefault(segs[0], set()).add(h)
            if len(segs) > 1:
                db.by_stem.setdefault("-".join(segs[:-1]), set()).add(h)
    return db


def lookup_hts(part: str, db: PartsDB) -> tuple[str, str, str]:
    """
    Resolve a part number to its HTS code.

    Returns (hts, how, note). `hts` is "" when nothing safe could be resolved,
    in which case `note` explains why -- the cell ships BLANK and the part is
    listed for review rather than being filled with a guess.

    Match order, most specific first:
      1. exact Product Key
      2. same key ignoring punctuation
      3. same key except the finish suffix   (1522807-U-BV -> 1522807-U-*)
      4. same leading part number            (1522807-U-BV -> 1522807-*)
    Steps 3 and 4 are only accepted when every candidate agrees on the code.
    """
    key = _norm_key(part)
    if key in db.by_key:
        return db.by_key[key], "exact", ""

    dp = _depunct(key)
    if dp in db.by_depunct and len(db.by_depunct[dp]) == 1:
        return next(iter(db.by_depunct[dp])), "punctuation-insensitive", ""

    segs = _segments(key)
    if segs:
        stem = "-".join(segs[:-1]) if len(segs) > 1 else ""
        if stem and stem in db.by_stem:
            codes = db.by_stem[stem]
            if len(codes) == 1:
                return next(iter(codes)), f"matched {stem}-* (finish suffix ignored)", ""
        base = segs[0]
        if base in db.by_base:
            codes = db.by_base[base]
            if len(codes) == 1:
                return next(iter(codes)), f"matched {base}-* (base part number)", ""
            return "", "conflict", (f"base {base} carries {len(codes)} different HTS codes "
                                    f"in the database ({', '.join(sorted(codes))}) and no "
                                    f"exact key exists -- resolve before filing")
    return "", "missing", "part number is not in the parts database"


# ---------------------------------------------------------------------------
# Join, allocate, aggregate
# ---------------------------------------------------------------------------
@dataclass
class ShipmentResult:
    rows: list[dict] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    review: list[dict] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    exact_groups: list[str] = field(default_factory=list)
    estimated_groups: list[str] = field(default_factory=list)


def join_by_bucket(lines: list[InvoiceLine], pl: PackingList,
                   res: ShipmentResult) -> dict[str, str]:
    """
    Reconcile the invoice against the packing list at the (PO, part) level and
    return {packing part number -> invoice part number} for any that differ.

    Matching per LINE is wrong: the two documents agree on what shipped, not on
    how it was written down. One supplier splits a part across four invoice
    lines and two packing rows; another prints one row per pallet. Both sides
    are therefore totalled per (PO, part) and compared there.

    Never joins by row position: documents do not share a row order (TRAP 2).
    Hand-keyed part numbers carry typos (TRAP 3), so an unmatched part is
    retried punctuation-blind and then at edit distance 1 -- accepted only when
    exactly one candidate fits, and every rewrite is reported.
    """
    po_on_packing = any(r.po for r in pl.rows)
    if not po_on_packing:
        res.notes.append("the packing list states no purchase order, so line items were "
                         "matched on part number alone")

    def bucket(po: str, part: str) -> tuple:
        return (po if po_on_packing else "", part)

    inv_qty: dict[tuple, float] = {}
    for ln in lines:
        k = bucket(ln.po, ln.part)
        inv_qty[k] = inv_qty.get(k, 0.0) + ln.qty
    pack_qty: dict[tuple, float] = {}
    for row in pl.rows:
        k = bucket(row.po, row.part)
        pack_qty[k] = pack_qty.get(k, 0.0) + row.qty

    alias: dict[str, str] = {}
    unmatched_pack = [k for k in pack_qty if k not in inv_qty]
    unmatched_inv = [k for k in inv_qty if k not in pack_qty]

    # Resolve the leftovers against each other by part number, tolerating the
    # punctuation and single-character typos that hand-keying introduces.
    for pk in list(unmatched_pack):
        hits = [ik for ik in unmatched_inv
                if (not po_on_packing or ik[0] == pk[0])
                and (_depunct(ik[1]) == _depunct(pk[1])
                     or _edit_distance_1(_depunct(ik[1]), _depunct(pk[1])))]
        if len(hits) == 1:
            alias[pk[1]] = hits[0][1]
            res.warnings.append(f"part number rewritten for matching: packing list "
                                f"'{pk[1]}' -> invoice '{hits[0][1]}'")
            inv_qty[pk] = inv_qty.pop(hits[0])
            unmatched_inv.remove(hits[0])
            unmatched_pack.remove(pk)

    for k in unmatched_inv:
        res.problems.append(f"invoice PO {k[0] or '(none)'} part {k[1]} "
                            f"({inv_qty[k]:,.0f} pcs): nothing on the packing list")
    for k in unmatched_pack:
        res.problems.append(f"packing list PO {k[0] or '(none)'} part {k[1]} "
                            f"({pack_qty[k]:,.0f} pcs): nothing on the invoices")
    for k in inv_qty:
        if k in pack_qty and abs(inv_qty[k] - pack_qty[k]) > 0.5:
            res.problems.append(
                f"PO {k[0] or '(none)'} part {k[1]}: invoice says {inv_qty[k]:,.0f} pcs, "
                f"packing list says {pack_qty[k]:,.0f}")
    return alias


def build_rows(lines: list[InvoiceLine], pl: PackingList | None,
               alias: dict[str, str], db: PartsDB,
               res: ShipmentResult) -> list[dict]:
    """Aggregate to one row per part, allocating weight inside each HS group.

    Weights and cartons are read from the PACKING rows directly rather than
    through the invoice lines. The packing list is the document that states
    what physically shipped, and it is free to break a part across a different
    number of rows than the invoice uses -- per pallet, per carton type, per
    container. Reading it on its own terms removes that mismatch entirely.
    """
    per_group: dict[int, dict[str, dict]] = {}
    per_part: dict[str, dict] = {}
    for i, ln in enumerate(lines):
        p = per_part.setdefault(ln.part, {"qty": 0.0, "value": 0.0,
                                          "countries": set(), "inv_hs": set()})
        p["qty"] += ln.qty
        p["value"] += ln.ext_total
        if ln.country_raw:
            p["countries"].add(ln.country_raw)
        if ln.invoice_hs:
            p["inv_hs"].add(ln.invoice_hs)


    # Piece counts per (weight group, part), straight off the packing list.
    if pl is not None:
        for row in pl.rows:
            part = alias.get(row.part, row.part)
            gp = per_group.setdefault(row.group, {}).setdefault(part, {"qty": 0.0})
            gp["qty"] += row.qty

    gross_by_part: dict[str, float] = {}
    net_by_part: dict[str, float] = {}
    if pl is not None:
        for g, parts in sorted(per_group.items()):
            names = list(parts)
            qtys = [parts[n]["qty"] for n in names]
            gross = allocate_largest_remainder(pl.group_gross.get(g, 0.0), qtys)
            for n, gv in zip(names, gross):
                gross_by_part[n] = round(gross_by_part.get(n, 0.0) + gv, 2)
            # A document that states no net weight leaves the column blank
            # rather than repeating gross or writing a zero it cannot support.
            group_net_total = pl.group_net.get(g)
            if group_net_total is not None:
                for n, nv in zip(names, allocate_largest_remainder(group_net_total, qtys)):
                    net_by_part[n] = round(net_by_part.get(n, 0.0) + nv, 2)
            label = (f"group {g + 1} ({pl.group_gross.get(g, 0.0):,.2f} kg gross, "
                     f"{len(names)} part(s))")
            (res.exact_groups if len(names) == 1 else res.estimated_groups).append(label)

    # True-up. Each group's split is exact within that group, but the group
    # totals are themselves rounded to the cent, so a shipment with fifty
    # single-row groups can drift a few cents from the figure printed on the
    # packing list. Push the residual onto the heaviest rows -- the same
    # largest-remainder principle, applied once at the file level.
    if pl is not None:
        for stated, by_part in ((pl.total_gross, gross_by_part),
                                (pl.total_net, net_by_part)):
            if stated is None or not by_part:
                continue
            drift = int(round(stated * 100)) - int(round(sum(by_part.values()) * 100))
            if drift == 0:
                continue
            order = sorted(by_part, key=lambda n: -by_part[n])
            step = 1 if drift > 0 else -1
            for n in (order * (abs(drift) // len(order) + 1))[:abs(drift)]:
                by_part[n] = round(by_part[n] + step / 100.0, 2)

    rows: list[dict] = []
    for part, agg in per_part.items():
        hts, how, note = lookup_hts(part, db)
        if not hts:
            res.review.append({"Part #": part, "Issue": how, "Detail": note})
            res.problems.append(f"{part}: no HTS ({note})")
        elif how != "exact":
            res.notes.append(f"{part}: HTS {hts} {how}")

        iso, ok = ("", True)
        if agg["countries"]:
            if len(agg["countries"]) > 1:
                res.problems.append(
                    f"{part}: invoices disagree on country of origin "
                    f"({', '.join(sorted(agg['countries']))})")
            iso, ok = _iso_country(sorted(agg["countries"])[0])
            if not ok:
                res.warnings.append(f"{part}: country '{iso}' has no ISO code mapping "
                                    f"-- written as printed")
        else:
            res.problems.append(f"{part}: no country of origin on the invoices")

        rows.append({
            "Part #": part,
            "Qty": int(agg["qty"]) if float(agg["qty"]).is_integer() else agg["qty"],
            "HTS": hts,
            "Value": round(agg["value"], 2),
            "Cartons": None,
            "Gross Weight (kg)": gross_by_part.get(part),
            "Net Weight (kg)": net_by_part.get(part),
            "Country Origin": iso,
            "Charges": CHARGES,
            "Zone": ZONE,
        })

    rows.sort(key=lambda r: r["Value"], reverse=True)

    # Value gets the same true-up as the weights: rounding each part's total to
    # the cent can leave the column a cent away from the invoice grand total,
    # and a customs file should tie to the invoice exactly.
    if rows:
        target = int(round(sum(l.ext_total for l in lines) * 100))
        drift = target - int(round(sum(r["Value"] for r in rows) * 100))
        if 0 < abs(drift) <= max(10, len(rows)):
            step = 1 if drift > 0 else -1
            for i in range(abs(drift)):
                r = rows[i % len(rows)]
                r["Value"] = round(r["Value"] + step / 100.0, 2)

    # Cartons. Most paperwork gives a shipment TOTAL only, in which case the
    # whole total sits on the first row: the column still ties out and nothing
    # is invented. Some templates do state a count per line, and those are used
    # as-is -- but only when EVERY matched row has one and they add up to the
    # shipment total, so a partly-filled column can never masquerade as exact.
    if pl is not None and rows:
        per_part: dict[str, float] = {}
        usable = bool(pl.rows) and all(r.cartons is not None for r in pl.rows)
        if usable:
            for row in pl.rows:
                part = alias.get(row.part, row.part)
                per_part[part] = per_part.get(part, 0.0) + row.cartons
            total = sum(per_part.values())
            if pl.cartons is not None and abs(total - pl.cartons) > 0.5:
                res.warnings.append(
                    f"per-row carton counts add to {total:,.0f} but the shipment total "
                    f"is {pl.cartons:,.0f}; falling back to the total on the first row")
                usable = False
        if usable and per_part:
            for r in rows:
                c = per_part.get(r["Part #"])
                if c is not None:
                    r["Cartons"] = int(c) if float(c).is_integer() else c
            res.notes.append(
                f"cartons: per-part counts taken from the {pl.carton_source}, "
                f"totalling {sum(per_part.values()):,.0f}.")
        elif pl.cartons is not None:
            rows[0]["Cartons"] = (int(pl.cartons) if float(pl.cartons).is_integer()
                                  else pl.cartons)
            res.notes.append(
                f"cartons: shipment total {pl.cartons:,.0f} placed on the first row "
                f"({rows[0]['Part #']}); source was the {pl.carton_source}. The "
                f"paperwork does not state cartons per part.")
    return rows


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------
def reconcile(doc: InvoiceDoc, pl: PackingList | None, rep: ReceptionReport | None,
              rows: list[dict], res: ShipmentResult) -> None:
    """Every check below treats a MISSING value as a failure, not a pass."""
    # qty x unit price == extended total, on every line
    for ln in doc.lines:
        if abs(round(ln.qty * ln.unit_price, 2) - ln.ext_total) > 0.02:
            res.problems.append(
                f"invoice {ln.invoice} part {ln.part}: {ln.qty:,.0f} x {ln.unit_price} "
                f"= {ln.qty * ln.unit_price:,.2f}, but the line reads {ln.ext_total:,.2f}")
    for msg in doc.page_total_mismatches:
        res.problems.append(msg)

    inv_value = round(sum(l.ext_total for l in doc.lines), 2)
    inv_qty = sum(l.qty for l in doc.lines)

    # ALL summary sets combined -- reading only the last one invents a
    # discrepancy that does not exist.
    if doc.summaries:
        by_invoice: dict[str, float] = {}
        for s in doc.summaries:
            by_invoice[s.invoice] = round(by_invoice.get(s.invoice, 0.0) + s.total, 2)
        summary_value = round(sum(by_invoice.values()), 2)
        if abs(summary_value - inv_value) > 0.01:
            res.problems.append(
                f"invoice pages total {inv_value:,.2f} but the {len(doc.summary_totals) or 1} "
                f"summary set(s) total {summary_value:,.2f}")
        if doc.summary_totals:
            printed = round(sum(doc.summary_totals.values()), 2)
            if abs(printed - summary_value) > 0.01:
                res.problems.append(
                    f"summary rows add to {summary_value:,.2f} but the printed "
                    f"Items Total(s) say {printed:,.2f}")
        parsed_invoices = {l.invoice for l in doc.lines}
        missing = set(by_invoice) - parsed_invoices
        extra = parsed_invoices - set(by_invoice)
        if missing:
            res.problems.append(f"{len(missing)} invoice(s) listed in a summary but not "
                                f"parsed from any page: {', '.join(sorted(missing)[:8])}")
        if extra:
            res.problems.append(f"{len(extra)} parsed invoice(s) absent from every summary: "
                                f"{', '.join(sorted(extra)[:8])}")
    else:
        res.warnings.append("no Invoice Summary pages found -- the invoice total could not "
                            "be checked against a roll-up")

    if pl is not None:
        pl_qty = sum(r.qty for r in pl.rows)
        if abs(pl_qty - inv_qty) > 0.001:
            res.problems.append(f"packing list qty {pl_qty:,.0f} != invoice qty {inv_qty:,.0f}")
        if pl.total_qty is not None and abs(pl.total_qty - pl_qty) > 0.001:
            res.problems.append(f"packing-list rows sum to {pl_qty:,.0f} but its total row "
                                f"says {pl.total_qty:,.0f}")
        group_gross = round(sum(pl.group_gross.values()), 2)
        group_net = round(sum(v for v in pl.group_net.values() if v is not None), 2)
        if pl.total_gross is None:
            res.problems.append("packing list states no total gross weight")
        elif abs(group_gross - pl.total_gross) > 0.01:
            res.problems.append(f"HS-group gross weights sum to {group_gross:,.2f} but the "
                                f"packing-list total is {pl.total_gross:,.2f}")
        if all(v is None for v in pl.group_net.values()):
            # Not a failure to report as a discrepancy -- the document simply
            # does not carry the figure. Say so plainly; the column is blank.
            res.warnings.append("the packing list states no net weight at all, so the "
                                "Net Weight column is blank for this shipment")
        elif pl.total_net is None:
            res.problems.append("packing list states no total net weight")
        elif abs(group_net - pl.total_net) > 0.01:
            res.problems.append(f"HS-group net weights sum to {group_net:,.2f} but the "
                                f"packing-list total is {pl.total_net:,.2f}")
        if pl.cartons is None:
            res.warnings.append("packing list states no carton count -- the Cartons column "
                                "is blank")
        for n in pl.notes:
            res.warnings.append(n)

        out_gross = round(sum(r["Gross Weight (kg)"] or 0 for r in rows), 2)
        out_net = round(sum(r["Net Weight (kg)"] or 0 for r in rows), 2)
        if pl.total_gross is not None and abs(out_gross - pl.total_gross) > 0.01:
            res.problems.append(f"output gross {out_gross:,.2f} != packing list "
                                f"{pl.total_gross:,.2f}")
        if pl.total_net is not None and abs(out_net - pl.total_net) > 0.01:
            res.problems.append(f"output net {out_net:,.2f} != packing list "
                                f"{pl.total_net:,.2f}")
    else:
        res.warnings.append("no packing list -- weights and cartons are blank, and the "
                            "invoice quantities could not be cross-checked")

    if rep is not None and rep.rows:
        rep_qty = sum(r.qty_physical for r in rep.rows)
        if abs(rep_qty - inv_qty) > 0.001:
            res.problems.append(f"reception report qty {rep_qty:,.0f} != invoice qty "
                                f"{inv_qty:,.0f}")
        inv_parts = {l.part for l in doc.lines}
        rep_parts = {r.part for r in rep.rows}
        for p in sorted(rep_parts - inv_parts):
            close = [q for q in inv_parts if _depunct(q) == _depunct(p)
                     or _edit_distance_1(_depunct(q), _depunct(p))]
            if len(close) == 1:
                res.warnings.append(f"reception report part '{p}' reads as '{close[0]}' on "
                                    f"the invoices (hand-keying typo)")
            else:
                res.problems.append(f"reception report lists part '{p}', absent from every "
                                    f"invoice")
        for p in sorted(inv_parts - rep_parts):
            close = [q for q in rep_parts if _depunct(q) == _depunct(p)
                     or _edit_distance_1(_depunct(q), _depunct(p))]
            if not close:
                res.problems.append(f"part '{p}' is on the invoices but not on the "
                                    f"reception report")
        if rep.gross_weight is not None and pl is not None and pl.total_gross is not None:
            if abs(rep.gross_weight - pl.total_gross) > 0.5:
                res.problems.append(f"reception report gross {rep.gross_weight:,.2f} != "
                                    f"packing list {pl.total_gross:,.2f}")
        # The Mexican fraccion and the US HTS are different code systems, so a
        # difference is expected -- surfaced once, as information, not an error.
        frac = {r.part: r.fraccion for r in rep.rows if r.fraccion}
        conflicts = sum(1 for r in rows if r["HTS"] and frac.get(r["Part #"])
                        and frac[r["Part #"]] != r["HTS"])
        if conflicts:
            res.notes.append(f"{conflicts} part(s) carry a Mexican fraccion on the reception "
                             f"report that differs from the US HTS -- expected; the HTS "
                             f"column is the database's US code.")
    elif rep is not None:
        res.warnings.append("reception report found but no line items could be read from it")

    for r in rows:
        if not r["HTS"]:
            continue
        if not (8 <= len(str(r["HTS"])) <= 12) or not str(r["HTS"]).isdigit():
            res.problems.append(f"{r['Part #']}: HTS '{r['HTS']}' is not a plausible "
                                f"8-12 digit code")
    for r in rows:
        if not r["Country Origin"]:
            res.problems.append(f"{r['Part #']}: country of origin is blank")


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------
def write_workbook(rows: list[dict], review: list[dict], out_path: Path) -> Path:
    """Write the customs file. If Excel has the workbook open the save fails
    (TRAP 8), so fall back to a timestamped name and say so rather than dying."""
    df = pd.DataFrame(rows, columns=OUT_COLUMNS)
    try:
        _save(df, review, out_path)
        return out_path
    except PermissionError:
        alt = out_path.with_name(
            f"{out_path.stem}_{datetime.now().strftime('%H%M%S')}{out_path.suffix}")
        print(f"  [WARN] {out_path.name} is open in Excel; writing {alt.name} instead.")
        _save(df, review, alt)
        return alt


def _save(df: pd.DataFrame, review: list[dict], path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Summary", index=False)
        if review:
            pd.DataFrame(review).to_excel(writer, sheet_name="Review", index=False)
    format_workbook(path, len(df))


def format_workbook(path: Path, n_rows: int) -> None:
    wb = load_workbook(path)
    ws = wb["Summary"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    widths = {"Part #": 18, "Qty": 10, "HTS": 14, "Value": 14, "Cartons": 10,
              "Gross Weight (kg)": 15, "Net Weight (kg)": 15,
              "Country Origin": 14, "Charges": 10, "Zone": 7}
    fmt = {"Qty": "#,##0", "Value": "#,##0.00", "Cartons": "#,##0",
           "Gross Weight (kg)": "#,##0.00", "Net Weight (kg)": "#,##0.00",
           "Charges": "0.00"}

    headers = [c.value for c in ws[1]]
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(name, 12)
        if name in fmt:
            for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = fmt[name]
        # HTS must be TEXT or Excel eats leading zeros and reformats long codes.
        if name == "HTS":
            for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, min_col=idx, max_col=idx):
                for cell in row:
                    cell.number_format = "@"

    total_row = n_rows + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        cell = ws.cell(row=total_row, column=idx)
        cell.font = Font(bold=True)
        # Qty, Value, Cartons and the weights sum. HTS is a code, not a number.
        if name in {"Qty", "Value", "Cartons", "Gross Weight (kg)", "Net Weight (kg)"}:
            cell.value = f"=SUM({letter}2:{letter}{n_rows + 1})"
            cell.number_format = fmt[name]
    ws.freeze_panes = "A2"

    if "Review" in wb.sheetnames:
        rv = wb["Review"]
        for cell in rv[1]:
            cell.font = header_font
            cell.fill = header_fill
        for idx in range(1, rv.max_column + 1):
            letter = get_column_letter(idx)
            longest = max((len(str(c.value)) for c in rv[letter] if c.value is not None),
                          default=12)
            rv.column_dimensions[letter].width = min(max(longest + 2, 14), 70)
    wb.save(path)


# ---------------------------------------------------------------------------
# Document classification and per-shipment driver
# ---------------------------------------------------------------------------
def classify(path: Path) -> str:
    """Identify a document by its CONTENT: filenames vary by client."""
    suffix = path.suffix.lower()
    # Formats this tool never reads. Container inspection reports and photos
    # travel with the paperwork; naming them keeps them out of the
    # "unrecognized" list, which is meant for files that SHOULD have parsed.
    if suffix in {".docx", ".doc", ".msg", ".eml", ".jpg", ".jpeg", ".png",
                  ".txt", ".htm", ".html", ".xml", ".csv"}:
        return "transport"
    if suffix == ".pdf":
        try:
            pages = extract_pdf_pages(path)
        except Exception:
            return "unknown"
        head = "\n".join(pages[:3])
        if re.search(r"REPORTE\s+DE\s+RECEPCION|RECONOCIMIENTO\s+PREVIO", head, re.I):
            return "reception"
        # Transport paperwork travels with every shipment and holds nothing this
        # tool needs. Recognize it so it is skipped quietly instead of being
        # reported as an unrecognized file the team has to look into.
        if re.search(r"BILL\s+OF\s+LADING|NON-NEGOTIABLE\s+WAYBILL|SEA\s*WAYBILL|"
                     r"FORWARDER'?S\s+CARGO\s+RECEIPT|TELEX\s+RELEASE|"
                     r"ARRIVAL\s+NOTICE|BOOKING\s+CONFIRMATION", head, re.I):
            return "transport"
        # A combined invoice+packing document (CIPL) is filed as invoices: the
        # invoice side is the one this tool cannot do without. process_shipment
        # then re-reads the same file for a packing table if none arrived
        # separately.
        if re.search(r"\bInvoice\b", head, re.I) and re.search(r"P\.?\s*O\.?\s*No", head, re.I):
            return "invoices"
        if re.search(r"Invoice\s+Summary", head, re.I):
            return "invoices"
        if re.search(r"PACKING\s*(?:[/&]\s*WEIGHT\s*)?LIST", head, re.I):
            return "packing"
        return "unknown"
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        try:
            xl = pd.ExcelFile(path, engine="xlrd" if suffix == ".xls" else "openpyxl")
            has_invoice = has_packing = False
            for sheet in xl.sheet_names:
                df = _read_sheet(path, sheet)
                flat = " ".join(str(v).lower() for v in df.head(40).values.ravel())
                if "product key" in flat and "hts" in flat:
                    return "partsdb"
                inv = _find_header(df, INV_COLUMN_TOKENS, {"po", "part", "qty", "amount"})
                if inv and "price" in inv[2]:
                    has_invoice = True
                pk = _find_header(df, PL_COLUMN_TOKENS, {"po", "part", "qty"})
                if pk and ({"gross", "net", "cartons"} & set(pk[2])):
                    has_packing = True
            # A workbook holding both is filed as invoices -- the invoice side is
            # the one the tool cannot do without. process_shipment re-reads the
            # same workbook for its packing sheet.
            if has_invoice:
                return "invoices"
            if has_packing:
                return "packing"
        except Exception:
            return "unknown"
    return "unknown"


def read_invoices(paths: list[Path]) -> InvoiceDoc:
    """Read commercial invoices in whichever form they arrived, merging every
    invoice document in the shipment. Suppliers routinely send one file per
    invoice, and the customs file covers the whole shipment."""
    merged = InvoiceDoc(path=paths[0])
    for path in paths:
        if path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}:
            doc = parse_invoices_excel(path)
        else:
            # The classic one-invoice-per-page layout first; suppliers who
            # instead print a single table fall through to the geometric reader.
            try:
                doc = parse_invoices(path)
                if not doc.lines:
                    raise RuntimeError("no line items in the page layout")
            except Exception:
                doc = parse_invoices_pdf_table(path)
        merged.lines.extend(doc.lines)
        merged.summaries.extend(doc.summaries)
        merged.summary_totals.update(doc.summary_totals)
        merged.page_total_mismatches.extend(doc.page_total_mismatches)
        merged.invoice_pages += doc.invoice_pages
        merged.summary_pages += doc.summary_pages
    return merged


def read_packing_list(path: Path, res: ShipmentResult) -> PackingList | None:
    """Read a packing list in whichever form it arrived. A PDF that cannot be
    parsed and validated is reported and dropped, not half-used: the run then
    continues without weights rather than with weights nobody can trust."""
    try:
        if path.suffix.lower() == ".pdf":
            return parse_packing_list_pdf(path)
        return parse_packing_list(path)
    except Exception as e:
        res.problems.append(f"packing list {path.name} could not be used: {e}")
        print(f"  [WARN] {e}")
        return None


def process_shipment(docs: dict[str, list[Path]], db: PartsDB,
                     out_dir: Path, label: str) -> tuple[int, str]:
    """Run one shipment end to end. Returns (problem count, message)."""
    res = ShipmentResult()

    doc = read_invoices(docs["invoices"])
    print(f"  invoices ....... {doc.invoice_pages} invoice page(s), {len(doc.lines)} line "
          f"item(s), {doc.summary_pages} summary page(s), "
          f"{len(doc.summary_totals)} summary set(s)")
    if not doc.lines:
        raise RuntimeError("no invoice line items could be parsed")

    pl = None
    for candidate in docs["packing"]:
        pl = read_packing_list(candidate, res)
        if pl is not None:
            break
    if pl is None:
        # No packing list of its own. A combined invoice+packing document
        # carries the table inside the invoice file -- a workbook with one sheet
        # each, or a CIPL PDF -- so look there before giving up on weights.
        for inv_path in docs["invoices"]:
            try:
                pl = (parse_packing_list(inv_path)
                      if inv_path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}
                      else parse_packing_list_pdf(inv_path))
                print(f"  [INFO] packing table read from {inv_path.name} "
                      f"(combined invoice + packing list)")
                break
            except Exception:
                pl = None
    if pl is not None:
        print(f"  packing list ... {len(pl.rows)} row(s), {len(pl.group_gross)} weight "
              f"group(s), cartons {pl.cartons if pl.cartons is not None else 'n/a'}")

    rep = None
    if docs["reception"]:
        rep = parse_reception_report(docs["reception"][0])
        print(f"  reception ...... {len(rep.rows)} row(s)"
              f"{', guide ' + rep.guide if rep.guide else ''}")

    alias = join_by_bucket(doc.lines, pl, res) if pl else {}
    rows = build_rows(doc.lines, pl, alias, db, res)
    reconcile(doc, pl, rep, rows, res)

    shipment = (pl.shipment_no if pl and pl.shipment_no
                else (rep.guide if rep and rep.guide else label))
    out_path = out_dir / f"customs_summary_{shipment}.xlsx"
    written = write_workbook(rows, res.review, out_path)

    print()
    for n in res.notes:
        print(f"  [NOTE] {n}")
    for w in res.warnings:
        print(f"  [WARN] {w}")
    for p in res.problems:
        print(f"  [CHECK] {p}")
    if res.estimated_groups or res.exact_groups:
        print()
        print("  Per-row weights are ALLOCATED within each HS group in proportion to")
        print("  piece count; column totals are exact, individual rows assume equal")
        print("  weight per piece.")
        for g in res.exact_groups:
            print(f"    exact (single part)  : {g}")
        for g in res.estimated_groups:
            print(f"    estimated (multi part): {g}")

    total_value = sum(r["Value"] for r in rows)
    total_qty = sum(r["Qty"] for r in rows)
    msg = (f"{len(doc.lines)} line items -> {len(rows)} parts, "
           f"{total_qty:,.0f} pcs, ${total_value:,.2f} -> {written.name}")
    return len(res.problems), msg


DOC_KINDS = ("invoices", "packing", "reception", "transport", "partsdb", "unknown")


def collect(folder: Path) -> dict[str, list[Path]]:
    """Sort a shipment folder's files by what they are.

    Byte-identical copies are dropped first: these arrive as email attachments
    and the same document routinely appears twice, once as 'X.pdf' and once as
    'X (1).pdf'. Two copies of one invoice would otherwise double every figure.
    """
    docs: dict[str, list[Path]] = {k: [] for k in DOC_KINDS}
    seen: dict[str, Path] = {}
    for p in sorted(folder.rglob("*")):
        if p.is_dir() or p.name.startswith("~$") or p.suffix.lower() == ".zip":
            continue
        try:
            digest = hashlib.md5(p.read_bytes()).hexdigest()
        except Exception:
            continue
        if digest in seen:
            continue
        seen[digest] = p
        docs[classify(p)].append(p)
    return docs


def expand_zips(in_dir: Path) -> list[Path]:
    """Extract any .zip dropped in the input folder, one folder per zip.

    Shipment paperwork arrives from the carrier as a single zip, so accepting
    them directly saves unpacking every one by hand. Files that were already
    extracted keep working exactly as before -- both forms are supported.
    """
    made: list[Path] = []
    seen: dict[str, Path] = {}
    for z in sorted(in_dir.glob("*.zip")):
        target = in_dir / f"_{z.stem}"
        try:
            # The same shipment's zip often arrives twice ('X.zip' and
            # 'X (1).zip'). Compare the CONTENTS, not the archive bytes: a
            # re-zipped copy holds the same documents but hashes differently.
            with zipfile.ZipFile(z) as probe:
                digest = hashlib.md5(
                    ";".join(sorted(f"{i.file_size}:{i.CRC}" for i in probe.infolist()
                                    if not i.is_dir())).encode()).hexdigest()
            if digest in seen:
                print(f"  [ZIP] {z.name} holds the same documents as "
                      f"{seen[digest].name} -- skipped")
                continue
            seen[digest] = z
            if not target.exists():
                target.mkdir(parents=True, exist_ok=True)
                with zipfile.ZipFile(z) as zf:
                    for member in zf.infolist():
                        if member.is_dir():
                            continue
                        # Flatten and sanitize: never let an archive path escape
                        # the folder we are extracting into.
                        name = Path(member.filename).name
                        if not name:
                            continue
                        with zf.open(member) as src, open(target / name, "wb") as dst:
                            dst.write(src.read())
                print(f"  [ZIP] {z.name} -> {target.name}\\")
            made.append(target)
        except Exception as e:
            print(f"  [WARN] could not open {z.name}: {e}")
    return made


def run(in_dir: Path, out_dir: Path, db_dir: Path) -> int:
    if not in_dir.exists():
        print(f"Creating input folder: {in_dir}")
        in_dir.mkdir(parents=True, exist_ok=True)
    if not db_dir.exists():
        db_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        db = load_parts_db(db_dir)
    except Exception as e:
        print(f"[ERROR] {e}")
        print(f"\nPut the latest 'Kohler Parts for Upload ....xlsx' in:")
        print(f"  {db_dir}")
        print("The tool always uses the most recently modified file there.")
        return 1
    print(f"Parts database: {db.path.name}  ({len(db.by_key):,} part numbers)")

    # Carrier paperwork usually arrives as one zip per shipment; unpack those
    # first. A shipment is then either the input folder itself, or one subfolder
    # per shipment (extracted here or put there by hand).
    expand_zips(in_dir)
    subfolders = [p for p in sorted(in_dir.iterdir()) if p.is_dir()]
    shipments = subfolders or [in_dir]

    processed = problems = 0
    for folder in shipments:
        label = folder.name if folder != in_dir else "shipment"
        docs = collect(folder)
        if not docs["invoices"]:
            if folder == in_dir and not any(docs[k] for k in docs):
                print(f"\nNo documents in {in_dir}. Drop one shipment's invoices, packing "
                      f"list and reception report there, then run again.")
                return 0
            # Say what WAS recognized, so it is obvious whether the other
            # documents were understood and only the invoices are missing.
            seen = [f"{len(docs[k])} {k}" for k in ("packing", "reception", "transport",
                                                    "unknown") if docs[k]]
            print(f"\n[SKIP] {label}: no commercial invoice found"
                  + (f" (found {', '.join(seen)})" if seen else ""))
            problems += 1
            continue

        print(f"\n{'=' * 62}\n{label}\n{'=' * 62}")
        for kind in ("invoices", "packing", "reception"):
            for p in docs[kind]:
                print(f"  [{kind.upper():9}] {p.name}")
        for p in docs["transport"]:
            print(f"  [TRANSPORT] {p.name}  (not needed for this file)")
        for p in docs["partsdb"]:
            print(f"  [PARTSDB  ] {p.name}  (ignored here -- the database\\ copy is used)")
        for p in docs["unknown"]:
            print(f"  [?????????] {p.name}  (not recognized -- check the file)")
            problems += 1
        for kind in ("invoices", "packing", "reception"):
            if len(docs[kind]) > 1:
                print(f"  [WARN] {len(docs[kind])} {kind} documents found; using "
                      f"{docs[kind][0].name}. Give each shipment its own subfolder "
                      f"under input\\ if these belong to different shipments.")
        print()

        try:
            n_problems, msg = process_shipment(docs, db, out_dir, label)
            print(f"\n  [OK  ] {msg}")
            if n_problems:
                print(f"  [!!!!] {n_problems} check(s) did not tie out -- read the "
                      f"[CHECK] lines above before filing.")
            problems += n_problems
            processed += 1
        except Exception as e:
            print(f"  [FAIL] {label}: {e}")
            traceback.print_exc()
            problems += 1

    print(f"\n{'=' * 62}")
    print(f"Done. {processed} shipment(s) processed.")
    print(f"Output folder: {out_dir}")
    print("=" * 62)
    return problems


def main() -> int:
    here = base_dir()
    in_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else here / "input"
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else here / "output"
    db_dir = Path(sys.argv[3]) if len(sys.argv) > 3 else here / "database"

    log_dir = here / "logs"
    log_dir.mkdir(exist_ok=True)
    log_path = log_dir / f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    tee = TeeLogger(log_path)
    sys.stdout = tee
    sys.stderr = tee

    print(f"Unimex Customs Summary  v{__version__}")
    print(f"Run started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Log file: {log_path}")
    print()

    problems = 0
    try:
        problems = run(in_dir, out_dir, db_dir)
    except Exception as e:
        print(f"\n[ERROR] Something went wrong: {e}")
        print("\nFull details (please send this log file to Andy):")
        traceback.print_exc()
        problems = 1  # a crash is also a reason to look for a newer build

    # Only check for updates when this run had trouble. A clean run never
    # touches the network. If a newer build installs it relaunches and
    # reprocesses the input, self-healing what this version couldn't handle.
    updated = False
    if getattr(sys, "frozen", False) and problems > 0:
        try:
            import updater  # lazy: only the frozen exe ever needs it
            updated = updater.check_and_update("customs", "UnimexCustoms.exe", __version__)
        except Exception as e:
            print(f"[update] skipped ({type(e).__name__}).")

    sys.stdout = tee.terminal
    sys.stderr = tee.terminal
    tee.close()
    if not updated:
        pause_for_user()  # on the update path the relaunched instance owns the UX
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
